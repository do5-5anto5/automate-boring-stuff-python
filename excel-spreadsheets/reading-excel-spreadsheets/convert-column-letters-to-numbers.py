import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# translate column number to a letter
b = get_column_letter(2)
aa = get_column_letter(27)
ahp = get_column_letter(900)
print(f'colums retrieved = {b}, {aa}, {ahp}')

worbook = openpyxl.load_workbook('excel-spreadsheets\example3.xlsx')

sheet = worbook['Sheet1']
max_column = get_column_letter(sheet.max_column)
print(f'{sheet.title} max column is {max_column}')


# get the number from given column letter
print(column_index_from_string('aa'))