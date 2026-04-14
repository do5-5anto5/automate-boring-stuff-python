import openpyxl
from pathlib import Path

# Create a Workbook object
workbook = openpyxl.Workbook()
# It starts off with a single sheet named 'Sheet'
print (workbook.sheetnames)

# Changing the sheet name
workbook.active.title = 'Spam Bacon Eggs'
print (workbook.sheetnames)

path = Path(__file__).parent

# Saving the workbook
workbook.save(path/'created_and_saved.xlsx')