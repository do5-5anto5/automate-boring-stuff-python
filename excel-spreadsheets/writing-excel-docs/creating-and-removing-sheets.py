import openpyxl

workbook = openpyxl.Workbook()

# Add a new sheet
workbook.create_sheet()
print(workbook.sheetnames)

# Create a new sheet, into a chosen index
workbook.create_sheet(index=0, title='First Sheet')
workbook.create_sheet(index=2, title='Middle Sheet')

print(workbook.sheetnames)

# Remove one sheet
del(workbook['Sheet'])
print(workbook.sheetnames)
# Remove many sheets
del(workbook['Middle Sheet'], workbook['Sheet1'])
print(workbook.sheetnames)