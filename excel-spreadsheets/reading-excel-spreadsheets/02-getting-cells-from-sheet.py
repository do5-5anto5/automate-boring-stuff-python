import openpyxl

wb_path = "excel-spreadsheets\\example3.xlsx"
workbook = openpyxl.load_workbook(wb_path)

sheet = workbook.active
a1 = sheet["A1"]  # get the cell

a1_value = a1.value  # get the value inside cell
print(a1_value)

c = sheet["B1"]
# get row, column and value from the cell
print(f"Row {c.row} Column {c.column} is {c.value}")

# get cell coordinates
print(f"Cell {c.coordinate} is {c.value}")

for i in range(1, 8, 2):
    print(sheet.cell(row=i, column=2).value)

max_row = sheet.max_row  # get sheet max row number
max_column = sheet.max_column  # get sheet max column number

print(
    f"""{sheet.title} 
max row number: {max_row}
max column number = {max_column}"""
)
